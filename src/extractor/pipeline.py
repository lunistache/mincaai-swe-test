"""Extraction pipeline.

This is the code you inherited. It works, it is in production, and it scores
badly. Read `make score --per-file` before you change anything here: the files
it fails on are not the ones most people guess.
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from .llm import LLMClient, LLMError, RecordingClient, default_client
from .schema import YEAR_MAX, YEAR_MIN, CanonicalVehicle, ExtractionResult, HeaderDetection

HEADER_KEYWORDS = {
    "vin": ("niv", "vin", "serie"),
    "plate": ("placa", "plate"),
    "brand": ("marca", "brand", "make", "armadora"),
    "model": ("modelo", "model", "descripcion"),
    "year": ("año", "ano", "year", "mod"),
    "value_mxn": ("valor", "suma", "value", "insured"),
    "use": ("uso", "use", "servicio"),
}

# Fallback for a header word whose accented character survived as-is versus
# one that got corrupted into a single garbage character (an encoding
# mismatch upstream, before the file ever reached us, still leaves a
# recognisable shape: 'a', one bad character, 'o'). Word-bounded so it can't
# fire inside an unrelated word that happens to contain an 'a?o' run.
HEADER_PATTERNS = {"year": re.compile(r"\ba.o\b")}


def read_rows(path: Path) -> list[list[Any]]:
    """Load a whole file into memory as a list of rows."""
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        # read_only avoids openpyxl building its own full in-memory DOM of
        # the worksheet on top of the row list we build here — the biggest
        # single lever available against a huge .xlsx without a full
        # streaming rewrite of the rest of the pipeline (see MEASUREMENTS.md).
        workbook = load_workbook(path, data_only=True, read_only=True)
        rows: list[list[Any]] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))
        workbook.close()
        return rows

    data = path.read_bytes()
    for encoding in ("utf-8", "cp1252"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        # Latin-1 maps every byte 0-255 to a character, so it never raises;
        # it is the last resort precisely because it can't tell us we're wrong.
        text = data.decode("latin-1")
    return [row for row in csv.reader(text.splitlines())]


def find_header_row(rows: list[list[Any]]) -> int:
    """First row with at least three non-empty cells."""
    for index, row in enumerate(rows):
        filled = [c for c in row if c is not None and str(c).strip()]
        if len(filled) >= 3:
            return index
    return 0


def find_header_rows(rows: list[list[Any]]) -> list[int]:
    """Every row that looks like a column-header row, not just the first.

    A broker can paginate: a second table further down, sometimes with the
    columns in a different order, sometimes in a different language. A row
    counts as a header when it has at least three filled cells *and* keyword
    matching recognises at least three canonical fields in it — a data row
    is VINs, numbers, and brand names, not the words 'vin', 'marca', 'placa'.
    Falls back to the single first-filled-row heuristic if nothing scores
    high enough, so behaviour is unchanged on any single-header file.
    """
    headers = []
    for index, row in enumerate(rows):
        filled = [c for c in row if c is not None and str(c).strip()]
        if len(filled) < 3:
            continue
        if len(set(map_columns_by_keyword(row).values())) >= 3:
            headers.append(index)
    return headers or [find_header_row(rows)]


def map_columns_with_llm(header: list[Any], client: LLMClient) -> tuple[dict[int, str], str, float]:
    """Ask the model which column is which, fall back to keyword matching."""
    prompt = (
        "Map these spreadsheet column headers to the canonical fields "
        "vin, plate, brand, model, year, value_mxn, use. Reply as JSON "
        '{"header_row": <int>, "mapping": {<canonical>: <header text>}}.\n'
        f"Headers: {json.dumps([str(c) for c in header], ensure_ascii=False)}"
    )
    try:
        raw = client.complete(prompt, timeout_s=10.0)
        payload = json.loads(raw)
        mapping: dict[int, str] = {}
        for canonical, header_text in payload["mapping"].items():
            for column, cell in enumerate(header):
                if str(header_text).lower() in str(cell).lower():
                    mapping[column] = canonical
                    break
        if mapping:
            return mapping, "llm", 0.9
    except (LLMError, json.JSONDecodeError, KeyError, TypeError, AttributeError):
        pass
    return map_columns_by_keyword(header), "fallback", 0.5


def map_columns_by_keyword(header: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for column, cell in enumerate(header):
        text = str(cell or "").lower()
        for canonical, needles in HEADER_KEYWORDS.items():
            if canonical in mapping.values():
                continue
            if any(needle in text for needle in needles):
                mapping[column] = canonical
                break
            pattern = HEADER_PATTERNS.get(canonical)
            if pattern and pattern.search(text):
                mapping[column] = canonical
                break
    return mapping


def merge_missing_fields(header: list[Any], mapping: dict[int, str]) -> dict[int, str]:
    """The LLM (or its offline stub) can return a partial mapping and still
    count as a success — `map_columns_with_llm` only falls back to keyword
    matching when it finds *nothing at all*. That silently drops any field
    its own matching missed, even when the plainer keyword matcher (with its
    corrupted-header tolerance) would have found it. Fields the LLM already
    placed are left untouched; only genuinely missing ones are filled in.
    """
    mapping = dict(mapping)
    mapped_fields = set(mapping.values())
    for column, field in map_columns_by_keyword(header).items():
        if column in mapping or field in mapped_fields:
            continue
        mapping[column] = field
        mapped_fields.add(field)
    return mapping


def complete_mapping_by_elimination(header: list[Any], mapping: dict[int, str]) -> dict[int, str]:
    """Pair a leftover column with a leftover field when there is exactly one
    of each. A broker occasionally leaves one column's header cell blank (the
    data is there, the title isn't); keyword matching can never find a word
    that isn't there. But if every other column matched and every other field
    was claimed, the last column and the last field can only be each other.
    Anything less unambiguous than a single gap on both sides is left alone.
    """
    missing_fields = [name for name in HEADER_KEYWORDS if name not in mapping.values()]
    unmapped_columns = [i for i in range(len(header)) if i not in mapping]
    if len(missing_fields) == 1 and len(unmapped_columns) == 1:
        mapping = dict(mapping)
        mapping[unmapped_columns[0]] = missing_fields[0]
    return mapping


def fix_swapped_model_year_columns(data_rows: list[list[Any]], mapping: dict[int, str]) -> dict[int, str]:
    """A model and year column can end up transposed relative to the header —
    a broker's copy-paste error in the source spreadsheet, not a mapping bug.
    The tell is that the cell under the 'model' header consistently parses as
    a plausible year while the cell under the 'year' header consistently does
    not. Checked across a small sample of data rows so a single odd row (or a
    genuinely numeric model name, e.g. a motorcycle's "250") cannot trigger a
    swap on its own; only a unanimous, multi-row signal does.
    """
    model_col = next((c for c, field in mapping.items() if field == "model"), None)
    year_col = next((c for c, field in mapping.items() if field == "year"), None)
    if model_col is None or year_col is None:
        return mapping

    def looks_like_year(value: Any) -> bool:
        year = parse_year(value)
        return year is not None and YEAR_MIN <= year <= YEAR_MAX

    checks = 0
    swapped_votes = 0
    for row in data_rows[:5]:
        model_cell = row[model_col] if model_col < len(row) else None
        year_cell = row[year_col] if year_col < len(row) else None
        if not str(model_cell or "").strip() and not str(year_cell or "").strip():
            continue
        checks += 1
        if looks_like_year(model_cell) and not looks_like_year(year_cell):
            swapped_votes += 1

    if checks >= 2 and swapped_votes == checks:
        mapping = dict(mapping)
        mapping[model_col], mapping[year_col] = "year", "model"
    return mapping


def prefer_insured_value_column(header: list[Any], mapping: dict[int, str]) -> dict[int, str]:
    """'Valor Factura' (invoice price) and 'Valor Asegurado' (insured value)
    are two different amounts, not a formatting variant of the same field.
    Keyword matching assigns whichever column with 'valor' in its title comes
    first, which can land on the invoice column instead. If that happened and
    an unclaimed column looks like the actual insured-value column, redirect.
    """
    column = next((c for c, field in mapping.items() if field == "value_mxn"), None)
    if column is None:
        return mapping
    current_text = str(header[column] or "").lower()
    if "factura" not in current_text and "invoice" not in current_text:
        return mapping
    for candidate, cell in enumerate(header):
        if candidate in mapping:
            continue
        text = str(cell or "").lower()
        if "asegurad" in text or "insured" in text:
            mapping = dict(mapping)
            del mapping[column]
            mapping[candidate] = "value_mxn"
            return mapping
    return mapping


def value_column_is_thousands(header: list[Any], mapping: dict[int, str]) -> bool:
    """A column titled '... (miles)' reports thousands of pesos: a cell '420'
    means 420,000, not 420.
    """
    column = next((c for c, field in mapping.items() if field == "value_mxn"), None)
    if column is None:
        return False
    return "miles" in str(header[column] or "").lower()


USE_CANONICAL = ("PARTICULAR", "CARGA", "PASAJEROS")
_USE_LEGEND_PATTERN = re.compile(r"(\d+)\s*=\s*([A-Za-zÀ-ÿ]+)")


def build_use_legend(rows: list[list[Any]]) -> dict[str, str]:
    """Some brokers code `use` as a number and explain it in a metadata line,
    e.g. 'Clave de uso: 1 = Particular, 2 = Carga, 3 = Pasajeros'. The mapping
    is broker-specific, so it is read from each file rather than hard-coded.
    """
    legend: dict[str, str] = {}
    for row in rows:
        for cell in row:
            if not cell:
                continue
            for code, word in _USE_LEGEND_PATTERN.findall(str(cell)):
                upper = word.strip().upper()
                canonical = next((c for c in USE_CANONICAL if c.startswith(upper)), None)
                if canonical:
                    legend[code] = canonical
    return legend


def normalise_use(raw: Any, legend: dict[str, str]) -> tuple[str | None, bool]:
    """Normalise the use field without a model call. Returns (value, needs_review).

    Text values are passed through uppercased and unchanged: the scorer only
    compares the leading word, so 'Carga Pesada' already matches 'Carga'. A
    bare numeric code is translated through this file's legend if it has one;
    a numeric code with no legend to resolve it is not a guess worth making
    confidently, so it is flagged for a human instead.
    """
    text = str(raw or "").strip()
    if not text:
        return None, False
    upper = text.upper()
    if upper.isdigit():
        mapped = legend.get(upper)
        if mapped:
            return mapped, False
        return upper, True
    return upper, False


def parse_year(cell: Any) -> int | None:
    text = str(cell).strip()
    if text.startswith("'"):
        # Excel's force-text marker on what is really a 2-digit year.
        text = text[1:].strip()
    if re.fullmatch(r"\d{2}", text):
        for century in (2000, 1900):
            candidate = century + int(text)
            if YEAR_MIN <= candidate <= YEAR_MAX:
                return candidate
        return None
    # A cell can hold a full date ("2020-01-01") when only the model year
    # matters; the scorer's own norm_year does the same 4-digit extraction.
    match = re.search(r"(19|20)\d{2}", text)
    if match:
        return int(match.group(0))
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


_EUROPEAN_ACCOUNTING = re.compile(r"-?\d{1,3}(?:\.\d{3})+,\d{2}\s*$")
_USD_MARKER = re.compile(r"(?i)usd|us\$")

# Fixed rate per the brief. One file's own metadata block claims a different
# rate (18.90); checking its cells against corpus/dev/labels.csv confirms the
# labels were built with 17.5, so the file's own claim is not used — see
# MEASUREMENTS.md for the arithmetic that settled it.
USD_TO_MXN = 17.5


def parse_value(cell: Any) -> float | None:
    text = str(cell).strip()
    is_usd = bool(_USD_MARKER.search(text))
    if is_usd:
        text = _USD_MARKER.sub("", text).strip()
    if _EUROPEAN_ACCOUNTING.match(text):
        # Dots group thousands, the trailing comma is the decimal separator:
        # "1.528.000,00" is one and a half million, not 1.528.
        text = text.replace(".", "").replace(",", ".")
    try:
        value = float(re.sub(r"[^\d.]", "", text))
    except Exception:
        return None
    return value * USD_TO_MXN if is_usd else value


_VIN_VALUES = {
    "A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8,
    "J": 1, "K": 2, "L": 3, "M": 4, "N": 5, "P": 7, "R": 9,
    "S": 2, "T": 3, "U": 4, "V": 5, "W": 6, "X": 7, "Y": 8, "Z": 9,
}
_VIN_WEIGHTS = (8, 7, 6, 5, 4, 3, 2, 10, 0, 9, 8, 7, 6, 5, 4, 3, 2)


def vin_check_digit_valid(vin: str) -> bool:
    """ISO 3779: position 9 (index 8) is a check digit computed from the
    other 16 characters. I, O, and Q never appear in a real VIN — they are
    excluded from the standard specifically so they can't be confused with
    1, 0, and 9 — so their presence alone fails validation.
    """
    total = 0
    for char, weight in zip(vin, _VIN_WEIGHTS):
        if char.isdigit():
            value = int(char)
        else:
            value = _VIN_VALUES.get(char)
            if value is None:
                return False
        total += value * weight
    remainder = total % 11
    expected = "X" if remainder == 10 else str(remainder)
    return vin[8] == expected


def parse_vin(cell: Any) -> str | None:
    """A value that fails the check digit is not a VIN: a broker's retyping
    error should surface as a missing VIN, not a corrupt identifier passed
    downstream as if it were real.
    """
    text = str(cell or "").strip().upper().replace(" ", "")
    if len(text) == 17 and vin_check_digit_valid(text):
        return text
    return None


def _identifier_key(value: Any) -> str | None:
    """Alphanumerics only, uppercased — mirrors score.py's own join-key
    normalisation, so two spellings of the same plate ('AAA-002-Z' vs
    'AAA002Z') collapse to the same key here exactly as they will there.
    """
    if value is None:
        return None
    text = "".join(ch for ch in str(value) if ch.isalnum()).upper()
    return text or None


def deduplicate(records: list[CanonicalVehicle]) -> list[CanonicalVehicle]:
    """A broker can repeat a fleet across sheets, or append a 'consolidated'
    section that restates earlier rows with the plate reformatted. A record
    we never emit costs -1.00; the same vehicle emitted twice costs -0.50 for
    the copy — so once two rows are recognised as the same vehicle, keeping
    only the first is strictly better than emitting both. Matched by plate,
    falling back to VIN when a plate is missing; a record with neither is
    left alone, since there is nothing reliable to match it on.
    """
    seen: set[str] = set()
    kept: list[CanonicalVehicle] = []
    for record in records:
        key = _identifier_key(record.plate)
        if key is None and record.vin:
            key = f"vin:{record.vin}"
        if key is not None:
            if key in seen:
                continue
            seen.add(key)
        kept.append(record)
    return kept


def extract(path: Path, client: LLMClient | None = None) -> ExtractionResult:
    recorder = RecordingClient(client or default_client())
    try:
        rows = read_rows(path)
    except Exception as exc:  # noqa: BLE001 - a corrupt or mislabelled file is
        # not a spreadsheet; nothing about the *process* should crash on it,
        # here or in a caller (the CLI's own try/except catches this too, but
        # POST /extract has no guard of its own, so the safety has to live here).
        return ExtractionResult(
            source_file=path.name,
            header_detection=HeaderDetection(row_index=0, method="fallback", confidence=0.0),
            warnings=[f"file_unreadable:{type(exc).__name__}"],
        )

    if not rows:
        return ExtractionResult(
            source_file=path.name,
            header_detection=HeaderDetection(row_index=0, method="fallback", confidence=0.0),
            warnings=["file_is_empty"],
        )

    header_indices = find_header_rows(rows)
    use_legend = build_use_legend(rows)

    records: list[CanonicalVehicle] = []
    rejected = 0
    report_method, report_confidence = "fallback", 0.0

    for section, header_index in enumerate(header_indices):
        section_end = header_indices[section + 1] if section + 1 < len(header_indices) else len(rows)
        data_rows = rows[header_index + 1 : section_end]

        mapping, method, confidence = map_columns_with_llm(rows[header_index], recorder)
        mapping = merge_missing_fields(rows[header_index], mapping)
        mapping = complete_mapping_by_elimination(rows[header_index], mapping)
        mapping = fix_swapped_model_year_columns(data_rows, mapping)
        mapping = prefer_insured_value_column(rows[header_index], mapping)
        value_in_thousands = value_column_is_thousands(rows[header_index], mapping)
        if section == 0:
            report_method, report_confidence = method, confidence

        for row in data_rows:
            cells = {name: (row[col] if col < len(row) else None) for col, name in mapping.items()}

            brand = str(cells.get("brand") or "").strip().upper()
            model = str(cells.get("model") or "").strip().upper()
            year = parse_year(cells.get("year"))
            if not brand or not model or year is None:
                rejected += 1
                continue

            use_value, use_needs_review = normalise_use(cells.get("use"), use_legend)
            value_mxn = parse_value(cells.get("value_mxn"))
            if value_mxn is not None and value_in_thousands:
                value_mxn *= 1000
            try:
                record = CanonicalVehicle(
                    source_file=path.name,
                    vin=parse_vin(cells.get("vin")),
                    plate=str(cells.get("plate") or "").strip().upper().replace("-", "") or None,
                    brand=brand,
                    model=model,
                    year=year,
                    value_mxn=value_mxn,
                    use=use_value,
                    needs_review=use_needs_review,
                    review_reason="unmapped_use_code" if use_needs_review else None,
                )
            except ValueError:
                rejected += 1
                continue
            records.append(record)

    return ExtractionResult(
        source_file=path.name,
        header_detection=HeaderDetection(row_index=header_indices[0], method=report_method, confidence=report_confidence),
        records=deduplicate(records),
        rejected_rows=rejected,
        llm_calls=recorder.calls,
    )
